# iterating through data

from pathlib import Path
from openpyxl import load_workbook
file_path = Path.cwd().joinpath('sample_data','reviews-sample.xlsx')
workbook = load_workbook(filename=file_path)
sheet = workbook.active

# You can slice the data with a combination of columns and rows:
for value in sheet.iter_rows(min_row=1,
                            max_row=2,
                            min_col=1,
                            max_col=3,
                            values_only=True):
    print(value)

# Print Column Names
for column_name in sheet.iter_rows(min_row=1,
                            max_row=1,
                            values_only=True):
    print(column_name)
# Alternatively This will print the first row if 1 is passed inside sheet[1] as parameter
for cell in sheet[1]:
    print(cell.value)

# Another way to iterate and edit the cell value
for cell in sheet['A1:G1']:
    for val in cell:
        if val.value == 'customer_id':
            val.value = 'cust_id'
        print(val.value)
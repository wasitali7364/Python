import pandas as pd

data = {
    "Product Name": ["Product 1", "Product 2"],
    "Sales Month 1": [10, 20],
    "Sales Month 2": [5, 35],
}
df = pd.DataFrame(data)
print(df)

# convert DataFrame into a worksheet

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path

workbook = Workbook()
sheet = workbook.active

for row in dataframe_to_rows(df, index=False, header=True):
    sheet.append(row)
    print(row)

workbook.save(Path.cwd().joinpath('sample_data','pandas.xlsx'))


# Convert worksheet into DataFrame

from openpyxl import load_workbook
workbook = load_workbook(Path.cwd().joinpath('sample_data','reviews-sample.xlsx'))
sheet = workbook.active

values = sheet.values

df = pd.DataFrame(values)

# get first row
header = df.iloc[0]

# remove first row from data
df = df[1:]

# set new header of df
df.columns = header

print(df.head())
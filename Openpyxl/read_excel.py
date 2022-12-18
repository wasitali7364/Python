# Reading Excel Spreadsheets With Openpyxl

from pathlib import Path
from openpyxl import load_workbook
file_path = Path.cwd().joinpath('sample_data','reviews-sample.xlsx')
workbook = load_workbook(filename=file_path)
'''
    *** read_only = True/False -> loads a spreadsheet in read-only mode allowing you to open very large Excel files.
    *** data_only = True/False -> ignores loading formulas and instead loads only the resulting values.
    *** Example:
                load_workbook(filename = file_path, read_only = True)
                load_workbook(filename = file_path, data_only = True)
'''

print(f'{workbook.sheetnames = }') # Print all the sheet names available in the excel workbook

sheet = workbook.active # selects first sheet (Sheet 1) by default

print(f'{sheet.title = }') # Prints the selected sheet name

print(f'{sheet["A1"].value = } is same as {sheet.cell(row=1, column=1).value = }')
print(f'{sheet["B1"].value = } is same as {sheet.cell(row=1, column=2).value = }')
'''
    *** Note: Even though in Python you are used to a zero-indexed notation,
    ***       with spreadsheets you will always use a one-indexed notation where the first row or column always has index 1.
'''
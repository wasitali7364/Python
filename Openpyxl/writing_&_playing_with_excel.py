# writing on excel sheet with openpyxl

from openpyxl import load_workbook
from pathlib import Path

file_path = Path.cwd().joinpath('sample_data','hello_world.xlsx')
workbook = load_workbook(file_path)
sheet = workbook.active

def print_rows():
    for row in sheet.iter_rows(values_only=True):
        print(row)

print_rows()

sheet['C1'] = 'Writing With Openpyxl!'

print_rows()
# above operations will not save any changes.

'''
Managing Rows and Columns
One of the most common things you have to do when manipulating spreadsheets is adding or removing rows and columns. The openpyxl package allows you to do that in a very straightforward way by using the methods:

.insert_rows()
.delete_rows()
.insert_cols()
.delete_cols()
Every single one of those methods can receive two arguments:

    1. idx
    2. amount

'''
# insert a column before the existing column 3 (C)
sheet.insert_cols(idx=3)
print_rows()

# insert 5 columns before the existing column 2 (B)
sheet.insert_cols(idx=2, amount= 5)
print_rows()

# delete created columns
sheet.delete_cols(idx=2, amount=5)
print_rows()

sheet.delete_cols(3)
print_rows()

# insert 3 row above row 1
sheet.insert_rows(1, amount= 3)
print_rows()

# delete the inserted row
sheet.delete_rows(1, 3)
print_rows()

'''
Change Sheet Name
'''
# Change Sheet Name
print(sheet.title) # output -> Sheet
# ---> Select the sheet whose name needs to change
selected_sheet = workbook['Sheet']
selected_sheet.title = 'My_New_Name'
print(selected_sheet.title) # output -> My_New_Name

# change sheet name back
selected_sheet.title = 'Sheet'

'''
Add or Delete Sheet

    ---> If you want to create, delete or copy sheets, then you can also do that with
        .create_sheet()
        .remove()
        .copy_worksheet()

'''
print(workbook.sheetnames)
new_sheet = workbook.create_sheet(title='My_New_Sheet')
print(workbook.sheetnames)

# Delete the newly created sheet
workbook.remove(new_sheet)  # Alternatively workbook.remove(workbook['My_New_Sheet'])
print(workbook.sheetnames)

# Create a duplicate copy of sheet
workbook.copy_worksheet(workbook['My_New_Name'])
print(workbook.sheetnames)

'''
    Freeze Panes with Openpyxl

        .freeze_panes = 'C2'
            --> notice that row 1 and columns A and B are frozen and are always visible no matter where you navigate within the spreadsheet.
'''
# Freeze Panes
sheet = workbook['My_New_Name Copy']
sheet.freeze_panes = 'C2'

# Check the used spreadsheet space using the attribute "dimensions"
print(sheet.dimensions)
sheet.auto_filter.ref = sheet.dimensions

'''
    Adding Formulas
'''
sheet['A5'] = 10
sheet['A6'] = 100
sheet['A7'] = '=SUM(A5:A6)'
print(sheet['A7'].value)

'''
    Adding Styles
        --> https://openpyxl.readthedocs.io/en/stable/styles.html
'''
# Import necessary style classes
from openpyxl.styles import Font, Color, Alignment, Border, Side

# Create a few styles
bold_font = Font(bold=True)
big_red_text = Font(color="00FF0000", size=20)
center_aligned_text = Alignment(horizontal="center")
double_border_side = Side(border_style="double")
square_border = Border(top=double_border_side,
                       right=double_border_side,
                       bottom=double_border_side,
                       left=double_border_side)

# Style some cells!
sheet["A1"].font = bold_font
sheet["B1"].font = big_red_text
sheet["C1"].alignment = center_aligned_text
sheet["A7"].border = square_border

# This will save the changes
workbook.save(file_path)